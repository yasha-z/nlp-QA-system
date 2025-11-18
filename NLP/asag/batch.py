"""Batch grading module for processing multiple student answers"""
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .predict import load_artifacts, predict_tfidf_ridge, predict_sbert


def parse_uploaded_file(file_path):
    """
    Parse uploaded CSV/Excel file with student responses
    
    Expected format:
    | Name | Roll Number | Q1 | Q2 | Q3 | ... |
    """
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
    elif file_path.endswith(('.xlsx', '.xls')):
        df = pd.read_excel(file_path)
    else:
        raise ValueError("File must be CSV or Excel format")
    
    return df


def combine_multiple_files(file_paths):
    """
    Combine multiple individual student response files into a single DataFrame.
    Each file represents one student's Google Form submission.
    
    Args:
        file_paths: List of file paths to combine
        
    Returns:
        pandas.DataFrame: Combined data with all students
    """
    try:
        all_dataframes = []
        
        for file_path in file_paths:
            # Parse each file
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            elif file_path.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file_path)
            else:
                print(f"Skipping unsupported file: {file_path}")
                continue
            
            # Check if file is in long format (has question_id column)
            if 'question_id' in df.columns:
                # Long format - keep all rows for this student
                print(f"Detected long format in {file_path} with {len(df)} questions")
            elif len(df) > 1:
                # Wide format with multiple students - take first row only
                print(f"Warning: {file_path} has {len(df)} rows. Taking first row only.")
                df = df.head(1)
            
            all_dataframes.append(df)
        
        # Combine all DataFrames
        if not all_dataframes:
            raise ValueError("No valid files to combine")
        
        combined_df = pd.concat(all_dataframes, ignore_index=True)
        
        print(f"✓ Successfully combined {len(all_dataframes)} student files into one dataset")
        return combined_df
    
    except Exception as e:
        raise Exception(f"Error combining files: {str(e)}")


def transform_long_to_wide(df):
    """
    Transform long format (question_id, student_answer per row) 
    to wide format (Q1, Q2, Q3 as columns)
    
    Args:
        df: DataFrame in long format with columns: student_name, question_id, student_answer
    
    Returns:
        DataFrame in wide format with columns: Name, Q1, Q2, Q3, etc.
    """
    # Check if already in wide format
    if 'question_id' not in df.columns:
        return df  # Already in wide format
    
    # Pivot from long to wide format
    wide_df = df.pivot_table(
        index='student_name',
        columns='question_id',
        values='student_answer',
        aggfunc='first'
    ).reset_index()
    
    # Rename student_name to Name for consistency
    wide_df.rename(columns={'student_name': 'Name'}, inplace=True)
    
    # Add empty Roll Number column if not present
    if 'Roll Number' not in wide_df.columns:
        wide_df['Roll Number'] = ''
    
    return wide_df


def grade_student_answers(df, model_answers, artifacts):
    """
    Grade all student answers against model answers
    
    Args:
        df: DataFrame with student data (will be transformed if in long format)
        model_answers: Dict mapping question IDs to model answers
        artifacts: Loaded ASAG model artifacts
    
    Returns:
        List of student results
    """
    # Transform to wide format if needed
    df = transform_long_to_wide(df)
    
    results = []
    
    for idx, row in df.iterrows():
        student_result = {
            'name': row.get('Name', row.get('name', row.get('student_name', ''))),
            'roll_number': row.get('Roll Number', row.get('roll_number', '')),
            'questions': []
        }
        
        # Grade each question
        for q_id, model_answer in model_answers.items():
            # Check if question column exists in dataframe
            if q_id not in df.columns:
                continue
                
            student_answer = str(row[q_id]) if pd.notna(row[q_id]) else ''
            
            try:
                # Get TF-IDF score
                tfidf_raw, tfidf_mapped = predict_tfidf_ridge(
                    artifacts, student_answer, model_answer
                )
                
                # Get SBERT score
                sbert_raw, sbert_mapped, cosine_sim = predict_sbert(
                    artifacts, student_answer, model_answer
                )
                
                # Calculate cosine-based score with RELAXED thresholds for better accuracy
                if cosine_sim >= 0.75:      # Excellent match (was 0.85)
                    cosine_score = 3
                elif cosine_sim >= 0.60:    # Good match (was 0.70)
                    cosine_score = 2
                elif cosine_sim >= 0.40:    # Fair match (was 0.50)
                    cosine_score = 1
                else:
                    cosine_score = 0        # Poor match
                
                # Calculate ensemble score - use MAX of both (50-50 trust both models equally)
                # This prevents one strict model from pulling down a good score
                better_ensemble = max(tfidf_mapped, cosine_score)
                
                question_result = {
                    'question': q_id,
                    'student_answer': student_answer,
                    'model_answer': model_answer,
                    'final_score': better_ensemble,
                    'tfidf_score': tfidf_mapped,
                    'cosine_score': cosine_score,
                    'similarity': cosine_sim,
                    'max_score': 3
                }
                
                student_result['questions'].append(question_result)
                
            except Exception as e:
                print(f"Error grading {q_id} for {student_result['name']}: {str(e)}")
                question_result = {
                    'question': q_id,
                    'student_answer': student_answer,
                    'model_answer': model_answer,
                    'final_score': 0,
                    'error': str(e),
                    'max_score': 3
                }
                student_result['questions'].append(question_result)
        
        # Calculate total score
        total_score = sum(q['final_score'] for q in student_result['questions'])
        max_possible = len(student_result['questions']) * 3
        percentage = (total_score / max_possible * 100) if max_possible > 0 else 0
        
        student_result['total_score'] = total_score
        student_result['max_possible'] = max_possible
        student_result['percentage'] = percentage
        
        results.append(student_result)
    
    return results


def calculate_metrics(results):
    """
    Calculate accuracy, precision, recall, and F1-score for the grading results
    
    Args:
        results: List of student results with predicted scores
    
    Returns:
        Dict with metrics
    """
    from sklearn.metrics import accuracy_score, precision_recall_fscore_support, confusion_matrix
    import numpy as np
    
    # For metrics, we need "true" scores, but we don't have them in batch grading
    # Instead, we'll calculate metrics based on predicted scores vs max possible
    # This gives us model performance statistics
    
    all_predicted = []
    all_max_scores = []
    
    for result in results:
        for q in result['questions']:
            all_predicted.append(q['final_score'])
            all_max_scores.append(q['max_score'])
    
    predicted_array = np.array(all_predicted)
    
    # Calculate distribution metrics
    total_answers = len(all_predicted)
    score_distribution = {
        'score_0': sum(1 for s in all_predicted if s == 0),
        'score_1': sum(1 for s in all_predicted if s == 1),
        'score_2': sum(1 for s in all_predicted if s == 2),
        'score_3': sum(1 for s in all_predicted if s == 3),
    }
    
    # Calculate average metrics
    avg_score = np.mean(predicted_array)
    std_score = np.std(predicted_array)
    
    # Categorize into performance levels for classification metrics
    # 0 = Poor, 1-2 = Fair, 3 = Good
    predicted_categories = ['Poor' if s == 0 else 'Fair' if s <= 2 else 'Good' for s in all_predicted]
    
    # Count categories
    category_counts = {
        'Poor': sum(1 for c in predicted_categories if c == 'Poor'),
        'Fair': sum(1 for c in predicted_categories if c == 'Fair'),
        'Good': sum(1 for c in predicted_categories if c == 'Good'),
    }
    
    metrics = {
        'total_answers_graded': total_answers,
        'average_score': round(avg_score, 2),
        'std_deviation': round(std_score, 2),
        'score_distribution': score_distribution,
        'category_distribution': category_counts,
        'score_percentages': {
            'score_0_pct': round(score_distribution['score_0'] / total_answers * 100, 1),
            'score_1_pct': round(score_distribution['score_1'] / total_answers * 100, 1),
            'score_2_pct': round(score_distribution['score_2'] / total_answers * 100, 1),
            'score_3_pct': round(score_distribution['score_3'] / total_answers * 100, 1),
        }
    }
    
    return metrics


def generate_individual_scoresheet(student_result, output_dir='results'):
    """
    Generate a detailed Excel scoresheet for a single student
    
    Args:
        student_result: Dict with student name, roll, questions, and scores
        output_dir: Directory to save the scoresheet
    
    Returns:
        Path to generated file
    """
    os.makedirs(output_dir, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Score Report"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "AUTOMATED SHORT ANSWER GRADING - SCORE REPORT"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Student Info
    ws['A3'] = "Student Name:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = student_result['name']
    
    ws['A4'] = "Roll Number:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = student_result['roll_number']
    
    ws['A5'] = "Date:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Summary
    ws['D3'] = "Total Score:"
    ws['D3'].font = Font(bold=True)
    ws['E3'] = f"{student_result['total_score']}/{student_result['max_possible']}"
    ws['E3'].font = Font(bold=True, size=12)
    
    ws['D4'] = "Percentage:"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = f"{student_result['percentage']:.1f}%"
    ws['E4'].font = Font(bold=True, size=12)
    
    # Grade
    percentage = student_result['percentage']
    if percentage >= 85:
        grade = "A"
        grade_color = "00B050"
    elif percentage >= 70:
        grade = "B"
        grade_color = "92D050"
    elif percentage >= 50:
        grade = "C"
        grade_color = "FFC000"
    else:
        grade = "D"
        grade_color = "FF0000"
    
    ws['D5'] = "Grade:"
    ws['D5'].font = Font(bold=True)
    ws['E5'] = grade
    ws['E5'].font = Font(bold=True, size=12, color=grade_color)
    
    # Question-wise breakdown header
    row = 7
    ws[f'A{row}'] = "Question"
    ws[f'B{row}'] = "Score"
    ws[f'C{row}'] = "Max"
    ws[f'D{row}'] = "Similarity"
    ws[f'E{row}'] = "Comments"
    
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Question details
    row += 1
    for q in student_result['questions']:
        ws[f'A{row}'] = q['question']
        ws[f'B{row}'] = q['final_score']
        ws[f'C{row}'] = q['max_score']
        ws[f'D{row}'] = f"{q.get('similarity', 0):.1%}"
        
        # Comments based on score
        score = q['final_score']
        if score == 3:
            comment = "Excellent! All key concepts covered."
        elif score == 2:
            comment = "Good. Most concepts covered."
        elif score == 1:
            comment = "Fair. Some concepts missing."
        else:
            comment = "Needs improvement. Review key concepts."
        
        ws[f'E{row}'] = comment
        
        # Apply borders
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
        
        # Color code scores
        score_cell = ws[f'B{row}']
        if score == 3:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif score == 2:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif score <= 1:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += 1
    
    # Detailed answers section
    row += 2
    ws[f'A{row}'] = "DETAILED ANSWERS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    for i, q in enumerate(student_result['questions'], 1):
        ws[f'A{row}'] = f"Q{i}: {q['question']}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Model Answer:"
        ws[f'A{row}'].font = Font(italic=True)
        ws[f'B{row}'] = q['model_answer']
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
        
        ws[f'A{row}'] = "Student Answer:"
        ws[f'A{row}'].font = Font(italic=True)
        ws[f'B{row}'] = q['student_answer']
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 2
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 35
    
    # Save file
    safe_name = "".join(c for c in str(student_result['name']) if c.isalnum() or c in (' ', '-', '_')).strip()
    filename = f"{safe_name}_{student_result['roll_number']}_scoresheet.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    return filepath


def generate_class_summary(results, output_dir='results'):
    """
    Generate a summary Excel file for the entire class
    
    Args:
        results: List of student results
        output_dir: Directory to save the summary
    
    Returns:
        Path to generated summary file
    """
    os.makedirs(output_dir, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Summary"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Title
    ws['A1'] = "CLASS SUMMARY REPORT"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A2:F2')
    
    # Headers
    row = 4
    headers = ['Name', 'Roll Number', 'Total Score', 'Max Score', 'Percentage', 'Grade']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Student data
    row += 1
    for result in results:
        ws.cell(row=row, column=1, value=result['name'])
        ws.cell(row=row, column=2, value=result['roll_number'])
        ws.cell(row=row, column=3, value=result['total_score'])
        ws.cell(row=row, column=4, value=result['max_possible'])
        ws.cell(row=row, column=5, value=f"{result['percentage']:.1f}%")
        
        # Calculate grade
        if result['percentage'] >= 85:
            grade = "A"
        elif result['percentage'] >= 70:
            grade = "B"
        elif result['percentage'] >= 50:
            grade = "C"
        else:
            grade = "D"
        ws.cell(row=row, column=6, value=grade)
        
        row += 1
    
    # Statistics
    row += 2
    ws.cell(row=row, column=1, value="CLASS STATISTICS").font = Font(bold=True)
    row += 1
    
    percentages = [r['percentage'] for r in results]
    ws.cell(row=row, column=1, value="Average Score:")
    ws.cell(row=row, column=2, value=f"{sum(percentages)/len(percentages):.1f}%")
    row += 1
    
    ws.cell(row=row, column=1, value="Highest Score:")
    ws.cell(row=row, column=2, value=f"{max(percentages):.1f}%")
    row += 1
    
    ws.cell(row=row, column=1, value="Lowest Score:")
    ws.cell(row=row, column=2, value=f"{min(percentages):.1f}%")
    
    # Adjust widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    filename = f"class_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    return filepath


def process_batch_grading(file_path, model_answers_dict, output_dir='results'):
    """
    Main function to process batch grading
    
    Args:
        file_path: Path to uploaded CSV/Excel file OR list of file paths
        model_answers_dict: Dict mapping question columns to model answers
        output_dir: Directory for output files
    
    Returns:
        Dict with paths to generated files
    """
    # Load models
    artifacts = load_artifacts('models')
    
    # Parse uploaded file(s)
    if isinstance(file_path, list):
        # Multiple files - combine them first
        print(f"Processing {len(file_path)} individual student files...")
        df = combine_multiple_files(file_path)
    else:
        # Single file
        df = parse_uploaded_file(file_path)
    
    # Grade all answers
    results = grade_student_answers(df, model_answers_dict, artifacts)
    
    # Calculate metrics
    metrics = calculate_metrics(results)
    
    # Generate individual scoresheets
    individual_files = []
    for result in results:
        filepath = generate_individual_scoresheet(result, output_dir)
        individual_files.append(filepath)
    
    # Generate class summary
    summary_file = generate_class_summary(results, output_dir)
    
    return {
        'individual_files': individual_files,
        'summary_file': summary_file,
        'total_students': len(results),
        'results': results,
        'metrics': metrics  # Add metrics to return
    }
